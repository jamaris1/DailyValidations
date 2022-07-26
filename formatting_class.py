import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import logging


'''
This script contains the classes that combine the source, target, and stage csv into one.
It will then format the format the combine excel and perform all calculations. 
Author: Jaime Amaris
Creation Date: 2022-07-22
Last Modification Date: 2022-07-22
'''


class base_format_class:

    def __init__(self):
        pass

    def del_refs(self,ws2):
        ws2.cell(1, 1).value = 'DW_SF_FIN_AP_AGING_TRANS_STMT_DEL'
        ws2.cell(2, 1).value = 'DW_SF_FIN_INVOICE_PAY_REQUIREMENT_DEL'
        ws2.cell(3, 1).value = 'DW_SF_FIN_FS_CTL_DETA_EXCLUDE_DEL'
        ws2.cell(4, 1).value = 'DW_SF_FIN_SEGMENT_GROUP_TYPE_DEL'
        ws2.cell(5, 1).value = 'DW_SF_FIN_SGMT_GRP_TYPEVAL_SEGVAL_DEL'
        ws2.cell(6, 1).value = 'DW_SF_FIN_SGMT_GRP_TYPE_VALUE_DEL'
        ws2.cell(7, 1).value = 'DW_SF_FIN_SECURITY_ROLE_OBJECT_DEL'
        ws2.cell(8, 1).value = 'DW_SF_FIN_INVOICE_EXCHANGE_RATE_DEL'
        ws2.cell(9, 1).value = 'DW_SF_FIN_DOCUMENT_DETAIL_SEGMENT_DEL'
        ws2.cell(10, 1).value = 'DW_SF_FIN_INVOICE_DIST_SEGMENT_DEL'
        ws2.cell(11, 1).value = 'DW_SF_FIN_INVOICE_DETAIL_DEL'
        ws2.cell(12, 1).value = 'DW_SF_FIN_FS_CONTROL_DETAIL_DEL'
        ws2.cell(13, 1).value = 'DW_SF_FIN_AP_AGING_TRANSACTION_DEL'
        ws2.cell(14, 1).value = 'DW_SF_FIN_AP_SUMMARY_DEL'
        ws2.cell(15, 1).value = 'DW_SF_FIN_AR_AGING_TRANSACTION_DEL'
        ws2.cell(16, 1).value = 'DW_SF_FIN_AP_SUMMARY_REPORT_DEL'
        ws2.cell(17, 1).value = 'DW_SF_FIN_AR_AGING_DEL'
        ws2.cell(18, 1).value = 'DW_SF_FIN_CASH_RECEIPT_DIST_DEL'
        ws2.cell(19, 1).value = 'DW_SF_FIN_CASH_RECEIPT_CHARGEBACK_DEL'
        ws2.cell(20, 1).value = 'DW_SF_FIN_CASH_RECEIPT_DEL'
        ws2.cell(21, 1).value = 'DW_SF_FIN_DOCUMENT_DETAIL_DEL'
        ws2.cell(22, 1).value = 'DW_SF_FIN_INVOICE_SPLIT_TERM_DEL'
        ws2.cell(23, 1).value = 'DW_SF_FIN_INVOICE_DEL'
        ws2.cell(24, 1).value = 'DW_SF_FIN_IFI_AP_INVOICE_DEL'
        ws2.cell(25, 1).value = 'DW_SF_FIN_INVOICE_DISTRIBUTION_DEL'

        ws2.cell(1, 2).value = 'DW_SF_FIN_AP_AGING_TRANS_STMT'
        ws2.cell(2, 2).value = 'DW_SF_FIN_INVOICE_PAY_REQUIREMENT'
        ws2.cell(3, 2).value = 'DW_SF_FIN_FS_CTL_DETA_EXCLUDE'
        ws2.cell(4, 2).value = 'DW_SF_FIN_SEGMENT_GROUP_TYPE'
        ws2.cell(5, 2).value = 'DW_SF_FIN_SGMT_GRP_TYPEVAL_SEGVAL'
        ws2.cell(6, 2).value = 'DW_SF_FIN_SGMT_GRP_TYPE_VALUE'
        ws2.cell(7, 2).value = 'DW_SF_FIN_SECURITY_ROLE_OBJECT'
        ws2.cell(8, 2).value = 'DW_SF_FIN_INVOICE_EXCHANGE_RATE'
        ws2.cell(9, 2).value = 'DW_SF_FIN_DOCUMENT_DETAIL_SEGMENT'
        ws2.cell(10, 2).value = 'DW_SF_FIN_INVOICE_DIST_SEGMENT'
        ws2.cell(11, 2).value = 'DW_SF_FIN_INVOICE_DETAIL'
        ws2.cell(12, 2).value = 'DW_SF_FIN_FS_CONTROL_DETAIL'
        ws2.cell(13, 2).value = 'DW_SF_FIN_AP_AGING_TRANSACTION'
        ws2.cell(14, 2).value = 'DW_SF_FIN_AP_SUMMARY'
        ws2.cell(15, 2).value = 'DW_SF_FIN_AR_AGING_TRANSACTION'
        ws2.cell(16, 2).value = 'DW_SF_FIN_AP_SUMMARY_REPORT'
        ws2.cell(17, 2).value = 'DW_SF_FIN_AR_AGING'
        ws2.cell(18, 2).value = 'DW_SF_FIN_CASH_RECEIPT_DIST'
        ws2.cell(19, 2).value = 'DW_SF_FIN_CASH_RECEIPT_CHARGEBACK'
        ws2.cell(20, 2).value = 'DW_SF_FIN_CASH_RECEIPT'
        ws2.cell(21, 2).value = 'DW_SF_FIN_DOCUMENT_DETAIL'
        ws2.cell(22, 2).value = 'DW_SF_FIN_INVOICE_SPLIT_TERM'
        ws2.cell(23, 2).value = 'DW_SF_FIN_INVOICE'
        ws2.cell(24, 2).value = 'DW_SF_FIN_IFI_AP_INVOICE'
        ws2.cell(25, 2).value = 'DW_SF_FIN_INVOICE_DISTRIBUTION'


    def setup_metadata(self):

        file1 = 'C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\fin_stage_query.csv'
        file2 = 'C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\fin_source_query.csv'
        file3 = 'C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\fin_target_query.csv'

        csv1 = pd.read_csv(file1)
        csv1.head()
        csv2 = pd.read_csv(file2)
        csv2.head()
        csv3 = pd.read_csv(file3)
        csv3.head()

        merged_data = csv3.merge(csv1, on='TABLE_NAME', how='left').merge(csv2, on='TABLE_NAME', how='left')

        #merged_data = pd.merge(map(pd.read_csv, [file1, file2, file3]), ignore_index=True)
        merged_data.to_excel('C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\comparison.xlsx')

        print(merged_data)

        wb = load_workbook('C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\comparison.xlsx')
        ws1 = wb.active
        ws1.insert_rows(1, 1)
        ws1.delete_cols(1, 1)
        ws1.merge_cells('B1:E1')
        ws1.merge_cells('F1:H1')

        ws2= wb.create_sheet("delete_refs")
        self.del_refs(ws2)

        #VLOODUP to insert table names into delete_refs sheet
        l = 0
        for i in ws2.iter_rows():
            table_name = i[0].value
            logging.info(table_name)
            for j in ws1.iter_rows():
                if j[0].value == table_name:
                    l = l + 1
                    ws2.cell(l,3).value = j[1].value
                    ws2.cell(l,4).value = j[2].value

        #VLOODUP to insert delete counts into sheet1
        l = 1
        for i in ws1.iter_rows():
            table_name = i[0].value
            logging.info(table_name)
            row_number = i[0].row

            for j in ws2.iter_rows():

                if j[1].value == table_name:
                    l = l + 1
                    ws1.cell(row_number,5).value = j[2].value




        ws1.cell(1,2).value = 'COUNTS'
        currentCell = ws1.cell(1, 2)  # or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='center')
        ws1.cell(1, 2).font = Font(bold=True)

        ws1.cell(1,6).value = 'COUNT VALIDATIONS'
        currentCell = ws1.cell(1, 6)  # or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='center')
        ws1.cell(1,6).font = Font(bold=True)





        ws1['E2'] = 'DELETE COUNT'
        ws1.cell(2,5).font = Font(bold=True)
        ws1['F2'] = 'SOURCE vs STAGE'
        ws1.cell(2,6).font = Font(bold=True)
        ws1['G2'] = 'STAGE vs TARGET'
        ws1.cell(2,7).font = Font(bold=True)
        ws1['H2'] = 'SOURCE vs TARGET'
        ws1.cell(2,8).font = Font(bold=True)


        #Inserts formulas to calculate source vs target, stage vs target, and stage vs source
        for i in range(3, 115):
            source = ws1.cell(row=i, column=4).value
            target = ws1.cell(row=i, column=2).value
            stage = ws1.cell(row=i, column=3).value
            sourceVSstage = int(source or 0) - int(stage or 0)
            stageVSTarget = int(stage or 0) - int(target or 0)
            sourceVStarget = int(source or 0) - int(target or 0)

            ws1.cell(row=i, column=6).value = sourceVSstage
            ws1.cell(row=i, column=7).value = stageVSTarget
            ws1.cell(row=i, column=8).value = sourceVStarget

        #Formats the font of the borders of the table.
        regular = Side(border_style="medium", color="000000")
        for c in ws1['A2:H2'][0]:
            c.border = Border(bottom=regular, top=regular, left=regular, right=regular)
        for c in ws1['A1:H1'][0]:
            c.border = Border(bottom=regular, top=regular, left=regular, right=regular)
        for c in ws1['A3:A115']+ws1['B3:B115']+ws1['C3:C115']+ws1['D3:D115']+ws1['E3:E115']+ws1['F3:F115']+ws1['G3:G115']+ws1['H3:H115']:
            c[0].border = Border(bottom=regular, top=regular, left=regular, right=regular)

        #Remove DEL TABLES at the end
        ws1.delete_rows(116, 27)



        wb.save('C:\\Users\\jamaris\\PycharmProjects\\DailyValidations\\files\\comparison.xlsx')





